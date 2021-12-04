from openpyxl import load_workbook

excel_file = "BEES PUBLICATIONS DATABASE.xlsx"

wb = load_workbook(filename=excel_file)
ws = wb.active

for row in ws.iter_rows(min_row=21, min_col=1, max_col=9, max_row=25):
    print(
        "<p>%s, \"%s\", <em>%s,</em> %s.<br><strong>DOI:&nbsp;</strong><a href=\"https://dx.doi.org/%s\">%s</a></p>" % (
            row[1].value, row[2].value, row[3].value, row[6].value, row[7].value, row[7].value))
